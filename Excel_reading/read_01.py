import openpyxl as xl
wb = xl.load_workbook('transations.xlsx')
sheet = wb['Sheet1']
cell =sheet['a1']
sell= sheet.cell(1,1)
print(cell.value)

for row in range(2,sheet.max_row+1):
    cell =sheet.cell(row,3)
    print(cell.value)
    corrected_price =cell.value*0.9
    corrected_price_cell =sheet.cell(row,4)
    corrected_price_cell.value=corrected_price
    
wb.save('transation_2.elsx')
